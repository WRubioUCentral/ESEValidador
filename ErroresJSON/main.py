"""
Validador de Archivos RIPS - Versión 2.0
Resoluciones 2275 de 2023 y 3280 de 2018

Script principal con validaciones avanzadas y corrección automática
"""
import os
import sys
import argparse
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
import io

# Agregar el directorio src al path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from src.file_reader import RIPSFileReader
from src.cross_validator import CrossFileValidator
from src.report_generator import ExcelReportGenerator
from src.logger import get_logger
from src.rules import ValidationError
from src.advanced_validators import DuplicateAttentionDetector, CIE10Validator, CoherenceValidator
from src.auto_corrector import AutoCorrector, apply_safe_corrections
from src.cie10_catalog import get_cie10_catalog


class ConsoleCapture:
    """Captura la salida de consola para incluirla en el informe"""

    def __init__(self):
        self.logs = []

    def write(self, message: str):
        """Escribe y captura el mensaje"""
        self.logs.append(message)
        sys.__stdout__.write(message)

    def get_logs(self) -> List[str]:
        """Obtiene todos los logs capturados"""
        return self.logs


class RIPSValidatorAppV2:
    """Aplicación principal del validador RIPS con funcionalidades avanzadas"""

    def __init__(self, input_dir: str = 'input', output_dir: str = 'output',
                 log_dir: str = 'logs', auto_correct: bool = False,
                 suggest_corrections: bool = False):
        """
        Inicializa la aplicación

        Args:
            input_dir: Directorio con archivos RIPS a validar
            output_dir: Directorio donde se guardará el informe
            log_dir: Directorio para logs
            auto_correct: Si True, aplica correcciones automáticas
            suggest_corrections: Si True, sugiere correcciones sin aplicarlas
        """
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.corrected_dir = os.path.join(output_dir, 'archivosCorregidos')
        self.log_dir = log_dir
        self.auto_correct = auto_correct
        self.suggest_corrections = suggest_corrections or auto_correct

        self.logger = get_logger(log_dir)
        self.console_capture = ConsoleCapture()

        # Componentes principales
        self.file_reader = RIPSFileReader()
        self.cross_validator = CrossFileValidator()

        # Validadores avanzados
        self.duplicate_detector = DuplicateAttentionDetector()
        self.cie10_validator = CIE10Validator()
        self.coherence_validator = CoherenceValidator()

        # Corrector automático
        self.auto_corrector = AutoCorrector(auto_mode=auto_correct)

        # Almacenamiento
        self.all_errors: List[ValidationError] = []
        self.corrected_files: Dict[str, List[str]] = {}  # file_type: [lines]

        self.stats = {
            'total_files': 0,
            'total_records': 0,
            'valid_records': 0,
            'invalid_records': 0,
            'files_by_type': {},
            'cie10_invalid_codes': 0,
            'duplicate_attentions': 0,
            'coherence_issues': 0,
            'corrections_applied': 0
        }

    def log_and_print(self, message: str, level: str = 'info'):
        """Log y print simultáneos con captura"""
        if level == 'info':
            self.logger.info(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)

        print(message)
        self.console_capture.write(message + '\n')

    def validate_input_directory(self) -> bool:
        """Valida que el directorio de entrada exista y contenga archivos"""
        if not os.path.exists(self.input_dir):
            self.log_and_print(f"ERROR: El directorio de entrada no existe: {self.input_dir}", 'error')
            return False

        txt_files = list(Path(self.input_dir).glob('*.txt'))
        if not txt_files:
            self.log_and_print(f"ADVERTENCIA: No se encontraron archivos .txt en: {self.input_dir}", 'warning')
            return False

        return True

    def collect_rips_files(self) -> Dict[str, List[str]]:
        """Recopila y clasifica archivos RIPS por tipo"""
        files_by_type = {
            'AF': [], 'US': [], 'AC': [], 'AP': [],
            'AT': [], 'AH': [], 'AM': [], 'AN': [], 'CT': []
        }

        for file_path in Path(self.input_dir).glob('*.txt'):
            file_type = self.file_reader.detect_file_type(file_path.name)
            if file_type:
                files_by_type[file_type].append(str(file_path))
                self.log_and_print(f"Archivo detectado: {file_path.name} -> Tipo: {file_type}")
            else:
                self.log_and_print(f"ADVERTENCIA: No se pudo detectar el tipo del archivo: {file_path.name}", 'warning')

        return files_by_type

    def process_file_with_advanced_validation(self, file_path: str, file_type: str) -> None:
        """
        Procesa un archivo con validaciones avanzadas y correcciones

        Args:
            file_path: Ruta al archivo
            file_type: Tipo de archivo RIPS
        """
        filename = os.path.basename(file_path)
        self.log_and_print(f"\n{'='*60}")
        self.log_and_print(f"Procesando: {filename}")
        self.log_and_print(f"{'='*60}")

        corrected_lines = []
        line_corrections = []

        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                for line_number, line in enumerate(f, start=1):
                    if not line.strip():
                        continue

                    original_line = line.strip()
                    corrected_line = original_line

                    # Aplicar correcciones si está habilitado
                    if self.auto_correct or self.suggest_corrections:
                        corrected_line, corrections = apply_safe_corrections(
                            original_line, line_number, filename, file_type, self.auto_corrector
                        )
                        line_corrections.extend(corrections)

                    corrected_lines.append(corrected_line)

                    # Validaciones específicas por tipo
                    fields = corrected_line.split(',')

                    # Validar CIE10 en archivos de atención
                    if file_type in ['AC', 'AP', 'AH']:
                        self._validate_cie10_in_record(fields, file_type, filename, line_number)

                    # Registrar atenciones para detección de duplicados
                    if file_type in ['AC', 'AP', 'AT', 'AH']:
                        self._register_attention(fields, file_type, filename, line_number)

                    # Validar coherencia
                    if file_type in ['AC', 'AP']:
                        self._validate_coherence_in_record(fields, file_type, filename, line_number)

            # Guardar archivo corregido si se aplicaron correcciones
            if self.auto_correct and line_corrections:
                self.corrected_files[filename] = corrected_lines
                self.stats['corrections_applied'] += len(line_corrections)
                self.log_and_print(f"  -> {len(line_corrections)} correcciones aplicadas")

            # Realizar validación estándar
            file_type_result, errors, file_stats = self.file_reader.read_and_validate_file(file_path)

            # Actualizar estadísticas
            self.stats['total_files'] += 1
            self.stats['total_records'] += file_stats.get('total_lines', 0)
            self.stats['valid_records'] += file_stats.get('valid_lines', 0)
            self.stats['invalid_records'] += file_stats.get('invalid_lines', 0)

            if file_type:
                self.stats['files_by_type'][file_type] = self.stats['files_by_type'].get(file_type, 0) + 1

            # Agregar errores
            self.all_errors.extend(errors)

            # Log de resultados
            if errors:
                self.log_and_print(f"ADVERTENCIA: Se encontraron {len(errors)} errores en {filename}", 'warning')
            else:
                self.log_and_print(f"OK: Archivo {filename} validado correctamente")

        except Exception as e:
            self.log_and_print(f"ERROR: Al procesar {filename}: {str(e)}", 'error')
            self.all_errors.append(ValidationError(
                filename, 0, "archivo",
                f"Error crítico al procesar el archivo: {str(e)}",
                "Sistema",
                "Verificar formato y estructura del archivo"
            ))

    def _validate_cie10_in_record(self, fields: List[str], file_type: str,
                                   filename: str, line_number: int):
        """Valida códigos CIE10 en un registro"""
        cie10_indices = []

        if file_type == 'AC':
            # AC: diagnostico_principal(11), relacionados(12-14)
            cie10_indices = [(11, 'diagnostico_principal', True),
                            (12, 'diagnostico_relacionado1', False),
                            (13, 'diagnostico_relacionado2', False),
                            (14, 'diagnostico_relacionado3', False)]
        elif file_type == 'AP':
            # AP: diagnostico_principal(12), relacionado(13), complicacion(14)
            cie10_indices = [(12, 'diagnostico_principal', True),
                            (13, 'diagnostico_relacionado', False),
                            (14, 'complicacion', False)]
        elif file_type == 'AH':
            # AH: diagnostico_ingreso(9), egreso(10), relacionados(11-13), complicacion(14)
            cie10_indices = [(9, 'diagnostico_ingreso', True),
                            (10, 'diagnostico_egreso', True),
                            (11, 'diagnostico_relacionado1', False),
                            (12, 'diagnostico_relacionado2', False),
                            (13, 'diagnostico_relacionado3', False),
                            (14, 'diagnostico_complicacion', False)]

        for idx, field_name, required in cie10_indices:
            if idx < len(fields):
                errors = self.cie10_validator.validate_code(
                    fields[idx], field_name, filename, line_number, required
                )
                if errors:
                    self.all_errors.extend(errors)
                    self.stats['cie10_invalid_codes'] += len(errors)

    def _register_attention(self, fields: List[str], file_type: str,
                           filename: str, line_number: int):
        """Registra una atención para detección de duplicados"""
        try:
            if file_type == 'AC' and len(fields) >= 7:
                # AC: tipo_doc(2), num_doc(3), fecha(4), cod_cups(6)
                self.duplicate_detector.register_attention(
                    fields[2], fields[3], fields[4], 'consulta', fields[6],
                    filename, line_number
                )
            elif file_type == 'AP' and len(fields) >= 7:
                # AP: tipo_doc(2), num_doc(3), fecha(4), cod_cups(6)
                self.duplicate_detector.register_attention(
                    fields[2], fields[3], fields[4], 'procedimiento', fields[6],
                    filename, line_number
                )
            elif file_type == 'AT' and len(fields) >= 7:
                # AT: tipo_doc(2), num_doc(3), cod_servicio(6)
                # Usar fecha del archivo AF correspondiente (simplificado)
                self.duplicate_detector.register_attention(
                    fields[2], fields[3], '', 'servicio', fields[6],
                    filename, line_number
                )
        except Exception as e:
            self.logger.debug(f"Error al registrar atención: {e}")

    def _validate_coherence_in_record(self, fields: List[str], file_type: str,
                                     filename: str, line_number: int):
        """Valida coherencia en un registro"""
        try:
            if file_type == 'AC' and len(fields) >= 22:
                # Coherencia finalidad-procedimiento-diagnóstico
                finalidad = fields[9]
                cod_cups = fields[6]
                diagnostico = fields[11]
                sexo = fields[21]
                edad = fields[19]
                unidad_edad = fields[20]

                # Validar coherencia
                errors = self.coherence_validator.validate_coherence(
                    finalidad, cod_cups, diagnostico, filename, line_number
                )
                self.all_errors.extend(errors)

                # Validar coherencia sexo-diagnóstico
                errors = self.coherence_validator.validate_gender_diagnosis_coherence(
                    sexo, diagnostico, filename, line_number
                )
                self.all_errors.extend(errors)

                # Validar coherencia edad-diagnóstico
                try:
                    edad_int = int(edad) if edad else 0
                    errors = self.coherence_validator.validate_age_diagnosis_coherence(
                        edad_int, unidad_edad, diagnostico, filename, line_number
                    )
                    self.all_errors.extend(errors)
                except ValueError:
                    pass

                if errors:
                    self.stats['coherence_issues'] += len(errors)

        except Exception as e:
            self.logger.debug(f"Error en validación de coherencia: {e}")

    def load_af_us_for_cross_validation(self, files_by_type: Dict[str, List[str]]):
        """Carga datos de AF y US para validaciones cruzadas"""
        self.log_and_print("\nCargando datos para validación cruzada...")

        # Cargar AF
        for af_file in files_by_type.get('AF', []):
            try:
                with open(af_file, 'r', encoding='utf-8', errors='replace') as f:
                    records = []
                    for line in f:
                        if line.strip():
                            fields = line.strip().split(',')
                            if len(fields) >= 5:
                                records.append({
                                    'num_factura': fields[4],
                                    'cod_prestador': fields[0]
                                })
                    self.cross_validator.register_af_data(records, os.path.basename(af_file))
                    self.log_and_print(f"  - Cargadas {len(records)} facturas de {os.path.basename(af_file)}")
            except Exception as e:
                self.log_and_print(f"ERROR: Al cargar AF: {e}", 'error')

        # Cargar US
        for us_file in files_by_type.get('US', []):
            try:
                with open(us_file, 'r', encoding='utf-8', errors='replace') as f:
                    records = []
                    for line in f:
                        if line.strip():
                            fields = line.strip().split(',')
                            if len(fields) >= 2:
                                records.append({
                                    'tipo_documento': fields[0],
                                    'num_documento': fields[1]
                                })
                    self.cross_validator.register_us_data(records, os.path.basename(us_file))
                    self.log_and_print(f"  - Cargados {len(records)} usuarios de {os.path.basename(us_file)}")
            except Exception as e:
                self.log_and_print(f"ERROR: Al cargar US: {e}", 'error')

    def perform_cross_validation(self, files_by_type: Dict[str, List[str]]):
        """Realiza validaciones cruzadas entre archivos"""
        self.log_and_print("\n" + "="*60)
        self.log_and_print("FASE 2: Validaciones cruzadas")
        self.log_and_print("="*60)

        # Validar referencias
        cross_validation_types = {
            'AC': self.cross_validator.validate_ac_references,
            'AP': self.cross_validator.validate_ap_references,
            'AT': self.cross_validator.validate_at_references,
            'AH': self.cross_validator.validate_ah_references,
            'AM': self.cross_validator.validate_am_references,
        }

        for file_type, validator_func in cross_validation_types.items():
            for file_path in files_by_type.get(file_type, []):
                try:
                    filename = os.path.basename(file_path)
                    self.log_and_print(f"Validando referencias en {filename}...")

                    with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                        records = []
                        for line in f:
                            if line.strip():
                                fields = line.strip().split(',')
                                if len(fields) >= 4:
                                    record = {
                                        'num_factura': fields[0],
                                        'tipo_documento': fields[2],
                                        'num_documento': fields[3]
                                    }
                                    records.append(record)

                    cross_errors = validator_func(records, filename)
                    if cross_errors:
                        self.log_and_print(f"  - {len(cross_errors)} errores de referencia", 'warning')
                        self.all_errors.extend(cross_errors)
                    else:
                        self.log_and_print(f"  - OK: Referencias correctas")

                except Exception as e:
                    self.log_and_print(f"ERROR: En validación cruzada de {filename}: {e}", 'error')

        # Verificar duplicados
        self.log_and_print("\nVerificando duplicados...")
        duplicate_errors = self.cross_validator.check_duplicates()
        if duplicate_errors:
            self.log_and_print(f"ADVERTENCIA: {len(duplicate_errors)} registros duplicados", 'warning')
            self.all_errors.extend(duplicate_errors)
        else:
            self.log_and_print("OK: No se encontraron duplicados")

    def detect_duplicate_attentions(self):
        """Detecta atenciones duplicadas"""
        self.log_and_print("\n" + "="*60)
        self.log_and_print("FASE 3: Detección de atenciones duplicadas")
        self.log_and_print("="*60)

        duplicate_errors = self.duplicate_detector.detect_duplicates()
        if duplicate_errors:
            self.stats['duplicate_attentions'] = len(duplicate_errors)
            self.log_and_print(f"ADVERTENCIA: {len(duplicate_errors)} atenciones duplicadas detectadas", 'warning')
            self.all_errors.extend(duplicate_errors)
        else:
            self.log_and_print("OK: No se detectaron atenciones duplicadas")

    def save_corrected_files(self):
        """Guarda los archivos corregidos"""
        if not self.corrected_files:
            return

        self.log_and_print("\n" + "="*60)
        self.log_and_print("Guardando archivos corregidos...")
        self.log_and_print("="*60)

        # Crear directorio
        os.makedirs(self.corrected_dir, exist_ok=True)

        for filename, lines in self.corrected_files.items():
            # Nombre del archivo corregido
            base_name = filename.replace('.txt', '')
            corrected_filename = f"{base_name}_corregido.txt"
            corrected_path = os.path.join(self.corrected_dir, corrected_filename)

            try:
                with open(corrected_path, 'w', encoding='utf-8') as f:
                    for line in lines:
                        f.write(line + '\n')

                self.log_and_print(f"  - Guardado: {corrected_filename}")
            except Exception as e:
                self.log_and_print(f"ERROR: Al guardar {corrected_filename}: {e}", 'error')

        self.log_and_print(f"\nArchivos corregidos guardados en: {self.corrected_dir}")

    def generate_comprehensive_report(self) -> str:
        """Genera el informe completo con todas las hojas"""
        self.log_and_print("\n" + "="*60)
        self.log_and_print("Generando informe completo...")
        self.log_and_print("="*60)

        output_file = os.path.join(self.output_dir, 'informe_errores.xlsx')
        report_generator = ExcelReportGenerator(output_file)

        # Estadísticas mejoradas
        enhanced_stats = {
            **self.stats,
            'cie10_validator_stats': {
                'invalid_codes': self.stats['cie10_invalid_codes'],
                'most_common_invalid': self.cie10_validator.get_most_common_invalid_codes(5)
            },
            'duplicate_stats': self.duplicate_detector.get_statistics(),
            'correction_stats': self.auto_corrector.get_corrections_summary()
        }

        # Generar informe base
        report_path = report_generator.generate_report(self.all_errors, enhanced_stats)

        # Agregar estadísticas de validación cruzada
        cross_stats = self.cross_validator.get_statistics()
        report_generator.add_cross_validation_summary(cross_stats)

        # Agregar hoja de correcciones
        if self.auto_corrector.corrections:
            corrections_export = self.auto_corrector.export_corrections()
            report_generator.add_corrections_sheet(corrections_export)
            self.log_and_print(f"  - Hoja 'Correcciones Realizadas': {len(corrections_export)} correcciones")

        # Agregar hoja de log de consola
        self._add_console_log_sheet(report_generator)

        # Guardar archivo final
        report_generator.workbook.save(output_file)

        self.log_and_print(f"\nInforme generado: {output_file}")
        return report_path

    def _add_console_log_sheet(self, report_generator: ExcelReportGenerator):
        """Agrega hoja con el log de la consola"""
        from openpyxl.styles import Font, Alignment

        log_sheet = report_generator.workbook.create_sheet(title="Log de Ejecución")

        # Título
        log_sheet.cell(row=1, column=1, value="Log de Ejecución del Validador").font = Font(bold=True, size=14)
        log_sheet.cell(row=2, column=1, value=f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # Logs
        logs = self.console_capture.get_logs()
        for idx, log_line in enumerate(logs, start=4):
            cell = log_sheet.cell(row=idx, column=1, value=log_line)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Ajustar ancho
        log_sheet.column_dimensions['A'].width = 150

        self.log_and_print(f"  - Hoja 'Log de Ejecución': {len(logs)} líneas")

    def print_summary(self):
        """Imprime resumen final"""
        self.log_and_print("\n" + "="*60)
        self.log_and_print("RESUMEN FINAL DE VALIDACIÓN")
        self.log_and_print("="*60)
        self.log_and_print(f"Archivos procesados: {self.stats['total_files']}")
        self.log_and_print(f"Registros procesados: {self.stats['total_records']}")
        self.log_and_print(f"Errores totales: {len(self.all_errors)}")
        self.log_and_print(f"  - Errores estándar: {len(self.all_errors) - self.stats['cie10_invalid_codes'] - self.stats['coherence_issues']}")
        self.log_and_print(f"  - Códigos CIE10 inválidos: {self.stats['cie10_invalid_codes']}")
        self.log_and_print(f"  - Problemas de coherencia: {self.stats['coherence_issues']}")
        self.log_and_print(f"  - Atenciones duplicadas: {self.stats['duplicate_attentions']}")

        if self.auto_correct:
            self.log_and_print(f"\nCorrección automática ACTIVADA:")
            self.log_and_print(f"  - Correcciones aplicadas: {self.stats['corrections_applied']}")
            summary = self.auto_corrector.get_corrections_summary()
            self.log_and_print(f"  - Alta confianza: {summary['alta_confianza']}")
            self.log_and_print(f"  - Requieren revisión: {summary['requieren_revision']}")

        self.log_and_print("="*60)

    def run(self):
        """Ejecuta el proceso completo de validación"""
        self.log_and_print("="*60)
        self.log_and_print("VALIDADOR DE ARCHIVOS RIPS - VERSIÓN 2.0")
        self.log_and_print("Resoluciones 2275 de 2023 y 3280 de 2018")
        if self.auto_correct:
            self.log_and_print("MODO: Corrección automática ACTIVADA")
        elif self.suggest_corrections:
            self.log_and_print("MODO: Sugerencia de correcciones")
        else:
            self.log_and_print("MODO: Solo validación")
        self.log_and_print("="*60)

        # Validar directorio
        if not self.validate_input_directory():
            return False

        # Recopilar archivos
        files_by_type = self.collect_rips_files()
        total_files = sum(len(files) for files in files_by_type.values())

        if total_files == 0:
            self.log_and_print("ADVERTENCIA: No se encontraron archivos RIPS válidos", 'warning')
            return False

        self.log_and_print(f"\nTotal de archivos RIPS encontrados: {total_files}")

        # FASE 1: Validación individual con validaciones avanzadas
        self.log_and_print("\n" + "="*60)
        self.log_and_print("FASE 1: Validación individual y avanzada")
        self.log_and_print("="*60)

        for file_type in ['AF', 'US', 'AC', 'AP', 'AT', 'AH', 'AM', 'AN', 'CT']:
            for file_path in files_by_type.get(file_type, []):
                self.process_file_with_advanced_validation(file_path, file_type)

        # Cargar datos para validación cruzada
        self.load_af_us_for_cross_validation(files_by_type)

        # FASE 2: Validaciones cruzadas
        self.perform_cross_validation(files_by_type)

        # FASE 3: Detección de duplicados de atenciones
        self.detect_duplicate_attentions()

        # Guardar archivos corregidos
        if self.auto_correct:
            self.save_corrected_files()

        # Generar informe completo
        report_path = self.generate_comprehensive_report()

        # Resumen final
        self.print_summary()

        if self.all_errors:
            self.log_and_print(f"\nADVERTENCIA: Se encontraron {len(self.all_errors)} errores en total", 'warning')
            self.log_and_print(f"Informe detallado: {report_path}")
        else:
            self.log_and_print("\nOK: Todos los archivos validados correctamente")

        return True


def main():
    """Función principal con argumentos CLI mejorados"""
    parser = argparse.ArgumentParser(
        description='Validador de archivos RIPS v2.0 - Con validaciones avanzadas y corrección automática',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Modos de operación:
  1. Solo validación (por defecto):
     python main_v2.py

  2. Validación con sugerencias de corrección:
     python main_v2.py --suggest-corrections

  3. Validación con corrección automática:
     python main_v2.py --auto-correct

Validaciones realizadas:
  - Validaciones estándar (formato, longitud, tipos de datos)
  - Códigos CIE10 contra catálogo vigente
  - Detección de atenciones duplicadas (mismo usuario, mismo día)
  - Coherencia entre finalidad, diagnóstico y procedimiento
  - Coherencia sexo-diagnóstico y edad-diagnóstico
  - Validaciones cruzadas entre archivos

Salidas generadas:
  - output/informe_errores.xlsx (informe completo)
  - output/archivosCorregidos/ (archivos corregidos, si se usa --auto-correct)
  - logs/ (logs detallados de ejecución)
        """
    )

    parser.add_argument(
        '-i', '--input',
        default='input',
        help='Directorio con archivos RIPS a validar (default: input)'
    )

    parser.add_argument(
        '-o', '--output',
        default='output',
        help='Directorio para guardar informes (default: output)'
    )

    parser.add_argument(
        '-l', '--logs',
        default='logs',
        help='Directorio para archivos de log (default: logs)'
    )

    parser.add_argument(
        '--auto-correct',
        action='store_true',
        help='Activa corrección automática de errores de alta confianza'
    )

    parser.add_argument(
        '--suggest-corrections',
        action='store_true',
        help='Sugiere correcciones sin aplicarlas (solo análisis)'
    )

    parser.add_argument(
        '--version',
        action='version',
        version='Validador RIPS v2.0.0'
    )

    args = parser.parse_args()

    # Validar combinación de argumentos
    if args.auto_correct and args.suggest_corrections:
        print("ADVERTENCIA: --auto-correct incluye --suggest-corrections. Se aplicarán correcciones automáticas.")

    # Crear y ejecutar validador
    try:
        validator = RIPSValidatorAppV2(
            input_dir=args.input,
            output_dir=args.output,
            log_dir=args.logs,
            auto_correct=args.auto_correct,
            suggest_corrections=args.suggest_corrections
        )
        success = validator.run()

        if success:
            print("\n[OK] Proceso completado exitosamente")
            sys.exit(0)
        else:
            print("\n[ERROR] El proceso finalizo con errores")
            sys.exit(1)

    except KeyboardInterrupt:
        print("\n\n[ADVERTENCIA] Proceso interrumpido por el usuario")
        sys.exit(130)
    except Exception as e:
        print(f"\n[ERROR] Error critico: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
