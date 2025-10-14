"""
Generador de informes de errores en formato Excel
"""
import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .rules import ValidationError


class ExcelReportGenerator:
    """Generador de informes Excel con formato profesional"""

    def __init__(self, output_path: str = None):
        """
        Inicializa el generador de informes

        Args:
            output_path: Ruta donde se guardará el archivo Excel
        """
        self.output_path = output_path or os.path.join('output', 'informe_errores.xlsx')
        self.workbook = Workbook()
        self.errors_sheet = self.workbook.active
        self.errors_sheet.title = "Errores Detectados"

    def _setup_header(self):
        """Configura el encabezado del informe con estilo"""
        headers = [
            'Nombre del Documento',
            'Número de Registro/Línea',
            'Nombre del Campo',
            'Descripción del Error',
            'Regla Normativa Asociada',
            'Corrección Recomendada'
        ]

        # Estilo del encabezado
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar encabezados
        for col_num, header in enumerate(headers, start=1):
            cell = self.errors_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        # Ajustar ancho de columnas
        column_widths = [30, 20, 30, 60, 30, 50]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col_num)
            self.errors_sheet.column_dimensions[column_letter].width = width

        # Congelar primera fila
        self.errors_sheet.freeze_panes = 'A2'

    def _add_error_row(self, row_num: int, error: ValidationError):
        """
        Agrega una fila con información de error

        Args:
            row_num: Número de fila
            error: Objeto ValidationError
        """
        data = [
            error.file_name,
            error.line_number,
            error.field_name,
            error.error_description,
            error.regulation,
            error.suggested_fix
        ]

        # Estilo para celdas de datos
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Alternar colores de fila
        if row_num % 2 == 0:
            data_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:
            data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Llenar celdas
        for col_num, value in enumerate(data, start=1):
            cell = self.errors_sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
            cell.fill = data_fill

    def generate_report(self, errors: List[ValidationError], stats: Dict[str, Any] = None) -> str:
        """
        Genera el informe completo de errores

        Args:
            errors: Lista de errores de validación
            stats: Estadísticas opcionales del proceso

        Returns:
            Ruta al archivo generado
        """
        # Configurar encabezado
        self._setup_header()

        # Agregar errores
        for idx, error in enumerate(errors, start=2):
            self._add_error_row(idx, error)

        # Crear hoja de resumen si hay estadísticas
        if stats:
            self._create_summary_sheet(errors, stats)

        # Guardar archivo
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self.workbook.save(self.output_path)

        return self.output_path

    def _create_summary_sheet(self, errors: List[ValidationError], stats: Dict[str, Any]):
        """
        Crea una hoja de resumen con estadísticas

        Args:
            errors: Lista de errores
            stats: Diccionario con estadísticas
        """
        summary_sheet = self.workbook.create_sheet(title="Resumen")

        # Título
        title_cell = summary_sheet.cell(row=1, column=1)
        title_cell.value = "Resumen de Validación RIPS"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='366092')
        summary_sheet.merge_cells('A1:B1')

        # Fecha de generación
        date_cell = summary_sheet.cell(row=2, column=1)
        date_cell.value = "Fecha de generación:"
        date_cell.font = Font(bold=True)
        summary_sheet.cell(row=2, column=2).value = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Estadísticas generales
        row = 4
        summary_sheet.cell(row=row, column=1, value="Estadísticas Generales").font = Font(bold=True, size=12)
        row += 1

        general_stats = [
            ("Total de archivos procesados", stats.get('total_files', 0)),
            ("Total de registros procesados", stats.get('total_records', 0)),
            ("Total de errores encontrados", len(errors)),
            ("Registros válidos", stats.get('valid_records', 0)),
            ("Registros con errores", stats.get('invalid_records', 0)),
        ]

        for label, value in general_stats:
            summary_sheet.cell(row=row, column=1, value=label)
            summary_sheet.cell(row=row, column=2, value=value)
            row += 1

        # Errores por archivo
        row += 2
        summary_sheet.cell(row=row, column=1, value="Errores por Archivo").font = Font(bold=True, size=12)
        row += 1

        # Contar errores por archivo
        errors_by_file: Dict[str, int] = {}
        for error in errors:
            errors_by_file[error.file_name] = errors_by_file.get(error.file_name, 0) + 1

        for filename, count in sorted(errors_by_file.items()):
            summary_sheet.cell(row=row, column=1, value=filename)
            summary_sheet.cell(row=row, column=2, value=count)
            row += 1

        # Tipos de error más comunes
        row += 2
        summary_sheet.cell(row=row, column=1, value="Campos con Más Errores").font = Font(bold=True, size=12)
        row += 1

        # Contar errores por campo
        errors_by_field: Dict[str, int] = {}
        for error in errors:
            errors_by_field[error.field_name] = errors_by_field.get(error.field_name, 0) + 1

        # Ordenar y tomar los top 10
        top_errors = sorted(errors_by_field.items(), key=lambda x: x[1], reverse=True)[:10]
        for field_name, count in top_errors:
            summary_sheet.cell(row=row, column=1, value=field_name)
            summary_sheet.cell(row=row, column=2, value=count)
            row += 1

        # Ajustar ancho de columnas
        summary_sheet.column_dimensions['A'].width = 40
        summary_sheet.column_dimensions['B'].width = 20

        # Aplicar bordes y alineación
        for row_cells in summary_sheet.iter_rows(min_row=4, max_row=row-1, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if cell.column == 2:
                    cell.alignment = Alignment(horizontal='right')

    def add_cross_validation_summary(self, cross_stats: Dict[str, int]):
        """
        Agrega una sección de resumen de validaciones cruzadas

        Args:
            cross_stats: Estadísticas de validación cruzada
        """
        if "Resumen" not in self.workbook.sheetnames:
            self._create_summary_sheet([], {})

        summary_sheet = self.workbook["Resumen"]

        # Encontrar la última fila con contenido
        last_row = summary_sheet.max_row + 3

        # Título de validación cruzada
        summary_sheet.cell(row=last_row, column=1, value="Validación Cruzada").font = Font(bold=True, size=12)
        last_row += 1

        cross_validation_stats = [
            ("Total de facturas en AF", cross_stats.get('total_facturas_af', 0)),
            ("Total de usuarios en US", cross_stats.get('total_usuarios_us', 0)),
            ("Facturas no encontradas", cross_stats.get('facturas_no_encontradas', 0)),
            ("Usuarios no encontrados", cross_stats.get('usuarios_no_encontrados', 0)),
            ("Registros duplicados", cross_stats.get('duplicados', 0)),
        ]

        for label, value in cross_validation_stats:
            summary_sheet.cell(row=last_row, column=1, value=label)
            summary_sheet.cell(row=last_row, column=2, value=value)

            # Aplicar bordes
            for col in [1, 2]:
                cell = summary_sheet.cell(row=last_row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            last_row += 1

    def add_corrections_sheet(self, corrections: List[Dict]):
        """
        Agrega una hoja con el registro de correcciones realizadas

        Args:
            corrections: Lista de correcciones (diccionarios)
        """
        if not corrections:
            return

        corrections_sheet = self.workbook.create_sheet(title="Correcciones Realizadas")

        # Encabezados
        headers = [
            'Nombre del Archivo',
            'Número de Línea',
            'Campo',
            'Valor Original',
            'Valor Corregido',
            'Tipo de Corrección',
            'Nivel de Confianza',
            'Razón',
            'Fecha y Hora'
        ]

        # Estilo del encabezado
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, start=1):
            cell = corrections_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        # Ajustar ancho de columnas
        column_widths = [25, 15, 25, 30, 30, 20, 18, 50, 20]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col_num)
            corrections_sheet.column_dimensions[column_letter].width = width

        # Agregar correcciones
        for idx, correction in enumerate(corrections, start=2):
            data = [
                correction.get('nombre_archivo', ''),
                correction.get('numero_linea', ''),
                correction.get('campo', ''),
                correction.get('valor_original', ''),
                correction.get('valor_corregido', ''),
                correction.get('tipo_correccion', ''),
                correction.get('confianza', ''),
                correction.get('razon', ''),
                correction.get('fecha_hora', '')
            ]

            # Estilo de datos
            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            data_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            # Color según nivel de confianza
            confidence = correction.get('confianza', '')
            if confidence == 'alta':
                data_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            elif confidence == 'media':
                data_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
            else:
                data_fill = PatternFill(start_color='FFCCBC', end_color='FFCCBC', fill_type='solid')

            for col_num, value in enumerate(data, start=1):
                cell = corrections_sheet.cell(row=idx, column=col_num)
                cell.value = value
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                cell.fill = data_fill

        # Congelar primera fila
        corrections_sheet.freeze_panes = 'A2'

        # Agregar leyenda de colores
        legend_row = len(corrections) + 4
        corrections_sheet.cell(row=legend_row, column=1, value="Leyenda de Confianza:").font = Font(bold=True)
        corrections_sheet.cell(row=legend_row+1, column=1, value="Verde").fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
        corrections_sheet.cell(row=legend_row+1, column=2, value="Alta confianza - Aplicada automáticamente")
        corrections_sheet.cell(row=legend_row+2, column=1, value="Amarillo").fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        corrections_sheet.cell(row=legend_row+2, column=2, value="Confianza media - Requiere revisión")
        corrections_sheet.cell(row=legend_row+3, column=1, value="Naranja").fill = PatternFill(start_color='FFCCBC', end_color='FFCCBC', fill_type='solid')
        corrections_sheet.cell(row=legend_row+3, column=2, value="Baja confianza - Requiere validación médica/administrativa")
