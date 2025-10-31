"""
Módulo para generar Excel con tablas especiales AC, AP, AH, AN, USUARIOS y RESUMEN
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Dict, List, Any
from src.cargador_rips import CargadorRIPS


class GeneradorTablasEspeciales:
    """Genera Excel con tablas especiales para análisis RIPS"""

    def __init__(self, cargador: CargadorRIPS):
        """
        Inicializa el generador

        Args:
            cargador: Instancia de CargadorRIPS
        """
        self.cargador = cargador

        # Estilos para encabezados
        self.header_font = Font(bold=True, color="FFFFFF", size=10)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Bordes
        thin_border = Side(style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        # Mapeos de códigos
        self.tipos_usuario = {
            "01": "Contributivo cotizante",
            "02": "Contributivo beneficiario",
            "03": "Contributivo adicional",
            "04": "Subsidiado",
            "05": "No asegurado",
            "06": "Especial o excepción cotizante",
            "07": "Especial o excepción beneficiario",
            "08": "Personas privadas de la libertad",
            "09": "Víctimas de eventos catastróficos"
        }

        self.regimenes = {
            "01": "CONTRIBUTIVO",
            "02": "CONTRIBUTIVO",
            "03": "CONTRIBUTIVO",
            "04": "SUBSIDIADO",
            "05": "NO ASEGURADO",
            "06": "ESPECIAL",
            "07": "ESPECIAL",
            "08": "ESPECIAL",
            "09": "ESPECIAL"
        }

        # Meses en español
        self.meses = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
            5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
            9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }

    def _aplicar_estilo_encabezado(self, ws, row_num: int, columnas: int):
        """Aplica estilo a la fila de encabezado"""
        for col in range(1, columnas + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

    def _aplicar_bordes(self, ws, start_row: int, end_row: int, columnas: int):
        """Aplica bordes a un rango de celdas"""
        for row in range(start_row, end_row + 1):
            for col in range(1, columnas + 1):
                ws.cell(row=row, column=col).border = self.border

    def _ajustar_columnas(self, ws):
        """Ajusta el ancho de las columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _obtener_regimen(self, tipo_usuario: str) -> str:
        """Obtiene el régimen a partir del tipo de usuario"""
        return self.regimenes.get(tipo_usuario, "DESCONOCIDO")

    def _formatear_fecha(self, fecha_str: str) -> str:
        """Convierte fecha de YYYY-MM-DD a DD/MM/YYYY"""
        try:
            if not fecha_str:
                return ""
            if " " in fecha_str:
                fecha_str = fecha_str.split(" ")[0]
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha.strftime("%d/%m/%Y")
        except:
            return fecha_str

    def _extraer_mes_nombre(self, fecha_str: str) -> str:
        """Extrae el nombre del mes en mayúsculas de una fecha"""
        try:
            if not fecha_str:
                return ""
            if " " in fecha_str:
                fecha_str = fecha_str.split(" ")[0]
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
            return self.meses.get(fecha.month, "")
        except:
            return ""

    def _calcular_edad_con_unidad(self, fecha_nacimiento: str, fecha_referencia: str = None):
        """
        Calcula edad y determina la unidad de medida apropiada

        Returns:
            tuple: (edad, unidad) donde unidad es 1=Años, 2=Meses, 3=Días
        """
        try:
            fn = datetime.strptime(fecha_nacimiento, "%Y-%m-%d")
            if fecha_referencia:
                if " " in fecha_referencia:
                    fecha_referencia = fecha_referencia.split(" ")[0]
                fr = datetime.strptime(fecha_referencia, "%Y-%m-%d")
            else:
                fr = datetime.now()

            # Calcular diferencia en días
            dias_total = (fr - fn).days

            if dias_total < 30:  # Menos de un mes
                return dias_total, "3"  # Días
            elif dias_total < 365:  # Menos de un año
                meses = dias_total // 30
                return meses, "2"  # Meses
            else:  # Un año o más
                edad = fr.year - fn.year
                if (fr.month, fr.day) < (fn.month, fn.day):
                    edad -= 1
                return edad, "1"  # Años
        except:
            return None, None

    def generar_tabla_ac_ap(self, datos_consolidados: List[Dict[str, Any]], nombre_archivo: str):
        """
        Genera Excel con tablas AC (Consultas) y AP (Procedimientos)

        Args:
            datos_consolidados: Lista con datos de todos los RIPS
            nombre_archivo: Nombre del archivo Excel a generar
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto

        # Tabla AC - Consultas (22 columnas exactas)
        ws_ac = wb.create_sheet("AC")
        encabezados_ac = [
            "Número de la factura",
            "Código del prestador de servicios de salud",
            "Tipo de identificación del usuario",
            "Número de identificación del usuario en el sistema",
            "Fecha del procedimiento",
            "Número de Autorización",
            "Código del procedimiento",
            "NOMBRE CUPS",
            "Ambito de realización del procedimiento",
            "Finalidad del procedimiento",
            "Personal que atiende",
            "Diagnóstivo principal",
            "Código del diagnóstico relacionado",
            "Complicación",
            "Forma de realización del acto quirúrgico",
            "Valor del Procedimiento",
            "EPS",
            "MUNICIPIO",
            "REGIMEN",
            "Edad",
            "Unidad de medida de la Edad",
            "SEXO",
            "MES"
        ]

        ws_ac.append(encabezados_ac)
        self._aplicar_estilo_encabezado(ws_ac, 1, len(encabezados_ac))

        row_ac = 2

        # Tabla AP - Procedimientos (22 columnas exactas)
        ws_ap = wb.create_sheet("AP")
        encabezados_ap = encabezados_ac.copy()

        ws_ap.append(encabezados_ap)
        self._aplicar_estilo_encabezado(ws_ap, 1, len(encabezados_ap))

        row_ap = 2

        # Procesar todos los RIPS
        for rips in datos_consolidados:
            datos = rips["datos"]
            num_factura = datos.get("numFactura", "")
            usuarios = self.cargador.extraer_usuarios(datos)

            for usuario in usuarios:
                tipo_doc = usuario.get("tipoDocumentoIdentificacion", "")
                num_doc = usuario.get("numDocumentoIdentificacion", "")
                tipo_usuario = usuario.get("tipoUsuario", "")
                fecha_nacimiento = usuario.get("fechaNacimiento", "")
                sexo = usuario.get("codSexo", "")
                cod_municipio = usuario.get("codMunicipioResidencia", "")
                regimen = self._obtener_regimen(tipo_usuario)

                # Procesar CONSULTAS para Tabla AC
                consultas = self.cargador.extraer_consultas(usuario)
                for consulta in consultas:
                    fecha_atencion = consulta.get("fechaInicioAtencion", "")
                    edad, unidad_edad = self._calcular_edad_con_unidad(fecha_nacimiento, fecha_atencion)
                    mes_nombre = self._extraer_mes_nombre(fecha_atencion)
                    fecha_formateada = self._formatear_fecha(fecha_atencion)

                    cod_consulta = consulta.get("codConsulta", "")
                    nombre_cups = self.cargador.obtener_nombre_cups(cod_consulta)

                    fila_ac = [
                        num_factura,
                        consulta.get("codPrestador", ""),
                        tipo_doc,
                        num_doc,
                        fecha_formateada,
                        consulta.get("numAutorizacion", ""),
                        cod_consulta,
                        nombre_cups,
                        consulta.get("modalidadGrupoServicioTecSal", ""),
                        consulta.get("finalidadTecnologiaSalud", ""),
                        "",  # Personal que atiende
                        consulta.get("codDiagnosticoPrincipal", ""),
                        consulta.get("codDiagnosticoRelacionado1", ""),
                        "",  # Complicación
                        "",  # Forma de realización
                        consulta.get("vrServicio", 0),
                        tipo_usuario,
                        cod_municipio,
                        regimen,
                        edad,
                        unidad_edad,
                        sexo,
                        mes_nombre
                    ]
                    ws_ac.append(fila_ac)
                    row_ac += 1

                # Procesar PROCEDIMIENTOS para Tabla AP
                procedimientos = self.cargador.extraer_procedimientos(usuario)
                for proc in procedimientos:
                    fecha_atencion = proc.get("fechaInicioAtencion", "")
                    edad, unidad_edad = self._calcular_edad_con_unidad(fecha_nacimiento, fecha_atencion)
                    mes_nombre = self._extraer_mes_nombre(fecha_atencion)
                    fecha_formateada = self._formatear_fecha(fecha_atencion)

                    cod_proc = proc.get("codProcedimiento", "")
                    nombre_cups = self.cargador.obtener_nombre_cups(cod_proc)

                    fila_ap = [
                        num_factura,
                        proc.get("codPrestador", ""),
                        tipo_doc,
                        num_doc,
                        fecha_formateada,
                        proc.get("numAutorizacion", ""),
                        cod_proc,
                        nombre_cups,
                        proc.get("modalidadGrupoServicioTecSal", ""),
                        proc.get("finalidadTecnologiaSalud", ""),
                        "",  # Personal que atiende
                        proc.get("codDiagnosticoPrincipal", ""),
                        proc.get("codDiagnosticoRelacionado", ""),
                        proc.get("codComplicacion", ""),
                        "",  # Forma de realización
                        proc.get("vrServicio", 0),
                        tipo_usuario,
                        cod_municipio,
                        regimen,
                        edad,
                        unidad_edad,
                        sexo,
                        mes_nombre
                    ]
                    ws_ap.append(fila_ap)
                    row_ap += 1

        # Aplicar bordes y ajustar columnas
        if row_ac > 2:
            self._aplicar_bordes(ws_ac, 2, row_ac - 1, len(encabezados_ac))
        self._ajustar_columnas(ws_ac)

        if row_ap > 2:
            self._aplicar_bordes(ws_ap, 2, row_ap - 1, len(encabezados_ap))
        self._ajustar_columnas(ws_ap)

        # Crear hojas AH, AN, USUARIOS, RESUMEN
        self._crear_tabla_ah(wb, datos_consolidados)
        self._crear_tabla_an(wb, datos_consolidados)
        self._crear_tabla_usuarios(wb, datos_consolidados)
        self._crear_tabla_resumen(wb, datos_consolidados)

        # Guardar archivo
        wb.save(nombre_archivo)
        print(f"\nExcel con tablas especiales generado: {nombre_archivo}")
        print(f"  - Tabla AC: {row_ac - 2} consultas")
        print(f"  - Tabla AP: {row_ap - 2} procedimientos")

    def _crear_tabla_ah(self, wb: Workbook, datos_consolidados: List[Dict[str, Any]]):
        """Crea la tabla AH - Hospitalizaciones (25 columnas exactas)"""
        ws = wb.create_sheet("AH")

        encabezados = [
            "FACTURA",
            "CODIGO MUNICIPIO",
            "TP DOC",
            "DOCUMENTO",
            "VIA DE INGRESO",
            "FECHA DE INGRESO",
            "HORA DE INGRESO",
            "AUTORIZACION",
            "CAUSA EXTERNA",
            "DIAGNOSTICO INGRESO",
            "DIAGNOSTICO EGRESO",
            "DIAGNOSTICO 2",
            "DIAGNOSTICO 3",
            "DIAGNOSTICO 4",
            "DIAGNOSTICO 5",
            "ESTADO AL EGRESO",
            "DX MUERTE",
            "FECHA EGRESO",
            "HORA EGRESO",
            "EPS",
            "MUNICIPIO",
            "REGIMEN",
            "Edad",
            "Unidad de medida de la Edad",
            "SEXO",
            "MES"
        ]

        ws.append(encabezados)
        self._aplicar_estilo_encabezado(ws, 1, len(encabezados))

        row = 2

        for rips in datos_consolidados:
            datos = rips["datos"]
            num_factura = datos.get("numFactura", "")
            usuarios = self.cargador.extraer_usuarios(datos)

            for usuario in usuarios:
                servicios = usuario.get("servicios", {})
                hospitalizaciones = servicios.get("hospitalizacion", [])

                if not hospitalizaciones:
                    continue

                tipo_doc = usuario.get("tipoDocumentoIdentificacion", "")
                num_doc = usuario.get("numDocumentoIdentificacion", "")
                tipo_usuario = usuario.get("tipoUsuario", "")
                fecha_nacimiento = usuario.get("fechaNacimiento", "")
                sexo = usuario.get("codSexo", "")
                cod_municipio = usuario.get("codMunicipioResidencia", "")
                regimen = self._obtener_regimen(tipo_usuario)

                for hosp in hospitalizaciones:
                    fecha_ingreso = hosp.get("fechaInicioAtencion", "")
                    edad, unidad_edad = self._calcular_edad_con_unidad(fecha_nacimiento, fecha_ingreso)
                    mes_nombre = self._extraer_mes_nombre(fecha_ingreso)

                    hora_ingreso = ""
                    if " " in fecha_ingreso:
                        hora_ingreso = fecha_ingreso.split(" ")[1] if len(fecha_ingreso.split(" ")) > 1 else ""

                    fecha_egreso = hosp.get("fechaEgreso", "")
                    hora_egreso = ""
                    if " " in fecha_egreso:
                        hora_egreso = fecha_egreso.split(" ")[1] if len(fecha_egreso.split(" ")) > 1 else ""

                    fila = [
                        num_factura,
                        cod_municipio,
                        tipo_doc,
                        num_doc,
                        hosp.get("viaIngresoServicioSalud", ""),
                        self._formatear_fecha(fecha_ingreso),
                        hora_ingreso,
                        hosp.get("numAutorizacion", ""),
                        hosp.get("causaMotivoAtencion", ""),
                        hosp.get("codDiagnosticoPrincipalIngreso", ""),
                        hosp.get("codDiagnosticoPrincipalEgreso", ""),
                        hosp.get("codDiagnosticoRelacionado1Egreso", ""),
                        hosp.get("codDiagnosticoRelacionado2Egreso", ""),
                        hosp.get("codDiagnosticoRelacionado3Egreso", ""),
                        "",  # Diagnostico 5
                        hosp.get("estadoSalidaPaciente", ""),
                        hosp.get("codDiagnosticoCausaMuerte", ""),
                        self._formatear_fecha(fecha_egreso),
                        hora_egreso,
                        tipo_usuario,
                        cod_municipio,
                        regimen,
                        edad,
                        unidad_edad,
                        sexo,
                        mes_nombre
                    ]
                    ws.append(fila)
                    row += 1

        if row > 2:
            self._aplicar_bordes(ws, 2, row - 1, len(encabezados))
        self._ajustar_columnas(ws)

    def _crear_tabla_an(self, wb: Workbook, datos_consolidados: List[Dict[str, Any]]):
        """Crea la tabla AN - Recién nacidos (17 columnas exactas)"""
        ws = wb.create_sheet("AN")

        encabezados = [
            "Número de la factura",
            "Código del prestador de servicios de salud",
            "Tipo de identificación del usuario",
            "Número de identificación del usuario en el sistema",
            "FECHA DE PARTO",
            "HORA DE O",
            "SEMANAS DE GESTACION",
            "ASISTIO A CONTROLES",
            "SEXO DEL RN",
            "PESO RN",
            "DX",
            "EPS",
            "MUNICIPIO",
            "REGIMEN",
            "Edad",
            "Unidad de medida de la Edad",
            "SEXO",
            "MES"
        ]

        ws.append(encabezados)
        self._aplicar_estilo_encabezado(ws, 1, len(encabezados))

        row = 2

        for rips in datos_consolidados:
            datos = rips["datos"]
            num_factura = datos.get("numFactura", "")
            usuarios = self.cargador.extraer_usuarios(datos)

            for usuario in usuarios:
                servicios = usuario.get("servicios", {})
                recien_nacidos = servicios.get("recienNacidos", [])

                if not recien_nacidos:
                    continue

                tipo_doc = usuario.get("tipoDocumentoIdentificacion", "")
                num_doc = usuario.get("numDocumentoIdentificacion", "")
                tipo_usuario = usuario.get("tipoUsuario", "")
                fecha_nacimiento = usuario.get("fechaNacimiento", "")
                sexo = usuario.get("codSexo", "")
                cod_municipio = usuario.get("codMunicipioResidencia", "")
                regimen = self._obtener_regimen(tipo_usuario)

                for rn in recien_nacidos:
                    fecha_parto = rn.get("fechaParto", "")
                    edad, unidad_edad = self._calcular_edad_con_unidad(fecha_nacimiento, fecha_parto)
                    mes_nombre = self._extraer_mes_nombre(fecha_parto)

                    hora_parto = ""
                    if " " in fecha_parto:
                        hora_parto = fecha_parto.split(" ")[1] if len(fecha_parto.split(" ")) > 1 else ""

                    fila = [
                        num_factura,
                        rn.get("codPrestador", ""),
                        tipo_doc,
                        num_doc,
                        self._formatear_fecha(fecha_parto),
                        hora_parto,
                        rn.get("edadGestacional", ""),
                        rn.get("consultasPrenatales", ""),
                        rn.get("sexo", ""),
                        rn.get("peso", ""),
                        rn.get("codDiagnosticoPrincipal", ""),
                        tipo_usuario,
                        cod_municipio,
                        regimen,
                        edad,
                        unidad_edad,
                        sexo,
                        mes_nombre
                    ]
                    ws.append(fila)
                    row += 1

        if row > 2:
            self._aplicar_bordes(ws, 2, row - 1, len(encabezados))
        self._ajustar_columnas(ws)

    def _crear_tabla_usuarios(self, wb: Workbook, datos_consolidados: List[Dict[str, Any]]):
        """Crea la tabla USUARIOS (18 columnas exactas)"""
        ws = wb.create_sheet("USUARIOS")

        encabezados = [
            "Tipo de Identificación del Usuario",
            "Número de Identifiación del Usuario en el Sistema",
            "Código Entidad Administradora",
            "Tipo de Usuario",
            "Primer Apellido del usuario",
            "Segundo apellido del usuario",
            "Primer nombre del usuario",
            "Segundo nombre del usuario",
            "Edad",
            "Unidad de medida de la Edad",
            "Sexo",
            "Código del departamento de residencia habitual",
            "Código de municipios de residencia habitual",
            "Zona de residencia habitual",
            "EPS",
            "MES",
            "REGIMEN",
            "SEDE"
        ]

        ws.append(encabezados)
        self._aplicar_estilo_encabezado(ws, 1, len(encabezados))

        row = 2
        usuarios_procesados = set()

        for rips in datos_consolidados:
            datos = rips["datos"]
            usuarios = self.cargador.extraer_usuarios(datos)

            for usuario in usuarios:
                tipo_doc = usuario.get("tipoDocumentoIdentificacion", "")
                num_doc = usuario.get("numDocumentoIdentificacion", "")

                clave_usuario = f"{tipo_doc}-{num_doc}"
                if clave_usuario in usuarios_procesados:
                    continue
                usuarios_procesados.add(clave_usuario)

                tipo_usuario = usuario.get("tipoUsuario", "")
                fecha_nacimiento = usuario.get("fechaNacimiento", "")
                edad, unidad_edad = self._calcular_edad_con_unidad(fecha_nacimiento)
                regimen = self._obtener_regimen(tipo_usuario)

                mes_nombre = ""
                servicios = usuario.get("servicios", {})
                consultas = servicios.get("consultas", [])
                if consultas:
                    mes_nombre = self._extraer_mes_nombre(consultas[0].get("fechaInicioAtencion", ""))

                cod_municipio = usuario.get("codMunicipioResidencia", "")
                cod_departamento = cod_municipio[:2] if len(cod_municipio) >= 2 else ""

                fila = [
                    tipo_doc,
                    num_doc,
                    "",  # Código Entidad Administradora
                    tipo_usuario,
                    usuario.get("primerApellido", ""),
                    usuario.get("segundoApellido", ""),
                    usuario.get("primerNombre", ""),
                    usuario.get("segundoNombre", ""),
                    edad,
                    unidad_edad,
                    usuario.get("codSexo", ""),
                    cod_departamento,
                    cod_municipio,
                    usuario.get("codZonaTerritorialResidencia", ""),
                    tipo_usuario,
                    mes_nombre,
                    regimen,
                    ""  # SEDE
                ]
                ws.append(fila)
                row += 1

        if row > 2:
            self._aplicar_bordes(ws, 2, row - 1, len(encabezados))
        self._ajustar_columnas(ws)

    def _crear_tabla_resumen(self, wb: Workbook, datos_consolidados: List[Dict[str, Any]]):
        """Crea la tabla RESUMEN con indicadores por usuario (43 columnas exactas)"""
        ws = wb.create_sheet("RESUMEN")

        encabezados = [
            "Tipo de Identificación del Usuario",
            "Número de Identifiación del Usuario en el Sistema",
            "Código Entidad Administradora",
            "Tipo de Usuario",
            "Primer Apellido del usuario",
            "Segundo apellido del usuario",
            "Primer nombre del usuario",
            "Segundo nombre del usuario",
            "Edad",
            "Unidad de medida de la Edad",
            "Sexo",
            "Código del departamento de residencia habitual",
            "Código de municipios de residencia habitual",
            "Zona de residencia habitual",
            "EPS",
            "MES",
            "REGIMEN",
            "SEDE",
            "CCU",
            "ODONTOLOGIA PRIMERA VEZ",
            "ODONTOLOGIA CONTROL",
            "PLACA BACTERIANA",
            "FLUOR",
            "DETARTRAJE",
            "SELLANTES",
            "VALORACIÓN INTEGRAL",
            "Primera Infancia",
            "Infancia",
            "Adolescencia",
            "Juventud",
            "Adultez",
            "Vejez",
            "AGUDEZA VISUAL",
            "PLANIFICACION FAMILIAR",
            "FECHA SUMINISTRO",
            "SUMINISTRO DE METODO",
            "SANGRE OCULTA",
            "COLESTEROL HDL",
            "COLESTEROL LDL",
            "COLESTEROL TOTAL",
            "GLICEMIA",
            "CREATININA",
            "TRIGLICERIDOS"
        ]

        ws.append(encabezados)
        self._aplicar_estilo_encabezado(ws, 1, len(encabezados))

        row = 2

        # Códigos CUPS para identificar procedimientos específicos
        codigos_cups = {
            "CCU": ["879101", "879102", "950601"],
            "ODONTOLOGIA_PRIMERA": ["890701"],
            "ODONTOLOGIA_CONTROL": ["890702"],
            "PLACA_BACTERIANA": ["997002", "997101"],
            "FLUOR": ["997106", "997310"],
            "DETARTRAJE": ["997001", "997302"],
            "SELLANTES": ["997101", "997311"],
            "VALORACION_INTEGRAL": ["890201", "890301"],
            "AGUDEZA_VISUAL": ["950101"],
            "PLANIFICACION": ["990205"],
        }

        codigos_lab = {
            "SANGRE_OCULTA": ["902222"],
            "COLESTEROL_HDL": ["903815"],
            "COLESTEROL_LDL": ["903816"],
            "COLESTEROL_TOTAL": ["903814"],
            "GLICEMIA": ["903841"],
            "CREATININA": ["903895"],
            "TRIGLICERIDOS": ["903818"]
        }

        usuarios_procesados = {}

        # Primera pasada: recolectar información de usuarios
        for rips in datos_consolidados:
            datos = rips["datos"]
            usuarios = self.cargador.extraer_usuarios(datos)

            for usuario in usuarios:
                tipo_doc = usuario.get("tipoDocumentoIdentificacion", "")
                num_doc = usuario.get("numDocumentoIdentificacion", "")
                clave_usuario = f"{tipo_doc}-{num_doc}"

                if clave_usuario not in usuarios_procesados:
                    tipo_usuario = usuario.get("tipoUsuario", "")
                    fecha_nacimiento = usuario.get("fechaNacimiento", "")
                    edad, unidad_edad = self._calcular_edad_con_unidad(fecha_nacimiento)
                    regimen = self._obtener_regimen(tipo_usuario)

                    curso_vida = self._determinar_curso_vida(edad, unidad_edad)

                    mes_nombre = ""
                    servicios = usuario.get("servicios", {})
                    consultas = servicios.get("consultas", [])
                    if consultas:
                        mes_nombre = self._extraer_mes_nombre(consultas[0].get("fechaInicioAtencion", ""))

                    cod_municipio = usuario.get("codMunicipioResidencia", "")
                    cod_departamento = cod_municipio[:2] if len(cod_municipio) >= 2 else ""

                    usuarios_procesados[clave_usuario] = {
                        "tipo_doc": tipo_doc,
                        "num_doc": num_doc,
                        "tipo_usuario": tipo_usuario,
                        "primer_apellido": usuario.get("primerApellido", ""),
                        "segundo_apellido": usuario.get("segundoApellido", ""),
                        "primer_nombre": usuario.get("primerNombre", ""),
                        "segundo_nombre": usuario.get("segundoNombre", ""),
                        "edad": edad,
                        "unidad_edad": unidad_edad,
                        "sexo": usuario.get("codSexo", ""),
                        "cod_departamento": cod_departamento,
                        "cod_municipio": cod_municipio,
                        "zona": usuario.get("codZonaTerritorialResidencia", ""),
                        "regimen": regimen,
                        "mes": mes_nombre,
                        "curso_vida": curso_vida,
                        "procedimientos": {},
                        "fechas_procedimientos": {}
                    }

                # Recolectar procedimientos con fechas
                procedimientos = self.cargador.extraer_procedimientos(usuario)
                for proc in procedimientos:
                    cod_proc = proc.get("codProcedimiento", "")
                    fecha_proc = proc.get("fechaInicioAtencion", "")
                    if cod_proc:
                        if cod_proc not in usuarios_procesados[clave_usuario]["procedimientos"]:
                            usuarios_procesados[clave_usuario]["procedimientos"][cod_proc] = 0
                            usuarios_procesados[clave_usuario]["fechas_procedimientos"][cod_proc] = fecha_proc
                        usuarios_procesados[clave_usuario]["procedimientos"][cod_proc] += 1

        # Segunda pasada: generar filas con formato especial para RESUMEN
        for clave, datos_usuario in usuarios_procesados.items():
            # Verificar procedimientos específicos y obtener fechas
            def obtener_fecha_procedimiento(codigos):
                for cod in codigos:
                    if cod in datos_usuario["procedimientos"]:
                        return self._formatear_fecha(datos_usuario["fechas_procedimientos"].get(cod, ""))
                return ""

            fecha_ccu = obtener_fecha_procedimiento(codigos_cups["CCU"])
            fecha_odonto_pv = obtener_fecha_procedimiento(codigos_cups["ODONTOLOGIA_PRIMERA"])
            fecha_odonto_ctrl = obtener_fecha_procedimiento(codigos_cups["ODONTOLOGIA_CONTROL"])
            fecha_placa = obtener_fecha_procedimiento(codigos_cups["PLACA_BACTERIANA"])
            fecha_fluor = obtener_fecha_procedimiento(codigos_cups["FLUOR"])
            fecha_detartraje = obtener_fecha_procedimiento(codigos_cups["DETARTRAJE"])
            fecha_sellantes = obtener_fecha_procedimiento(codigos_cups["SELLANTES"])
            fecha_valoracion = obtener_fecha_procedimiento(codigos_cups["VALORACION_INTEGRAL"])
            fecha_agudeza = obtener_fecha_procedimiento(codigos_cups["AGUDEZA_VISUAL"])
            fecha_planif = obtener_fecha_procedimiento(codigos_cups["PLANIFICACION"])

            # Laboratorios
            fecha_sangre = obtener_fecha_procedimiento(codigos_lab["SANGRE_OCULTA"])
            fecha_hdl = obtener_fecha_procedimiento(codigos_lab["COLESTEROL_HDL"])
            fecha_ldl = obtener_fecha_procedimiento(codigos_lab["COLESTEROL_LDL"])
            fecha_col_total = obtener_fecha_procedimiento(codigos_lab["COLESTEROL_TOTAL"])
            fecha_glicemia = obtener_fecha_procedimiento(codigos_lab["GLICEMIA"])
            fecha_creatinina = obtener_fecha_procedimiento(codigos_lab["CREATININA"])
            fecha_trigli = obtener_fecha_procedimiento(codigos_lab["TRIGLICERIDOS"])

            fila = [
                datos_usuario["tipo_doc"],
                datos_usuario["num_doc"],
                "",  # Código Entidad Administradora
                datos_usuario["tipo_usuario"],
                datos_usuario["primer_apellido"],
                datos_usuario["segundo_apellido"],
                datos_usuario["primer_nombre"],
                datos_usuario["segundo_nombre"],
                datos_usuario["edad"],
                datos_usuario["unidad_edad"],
                datos_usuario["sexo"],
                datos_usuario["cod_departamento"],
                datos_usuario["cod_municipio"],
                datos_usuario["zona"],
                datos_usuario["tipo_usuario"],
                datos_usuario["mes"],
                datos_usuario["regimen"],
                "",  # SEDE
                fecha_ccu,  # CCU - fecha si SI, vacío si NO
                fecha_odonto_pv,
                fecha_odonto_ctrl,
                fecha_placa,
                fecha_fluor,
                fecha_detartraje,
                fecha_sellantes,
                fecha_valoracion,
                "SI" if datos_usuario["curso_vida"] == "Primera Infancia" else "",
                "SI" if datos_usuario["curso_vida"] == "Infancia" else "",
                "SI" if datos_usuario["curso_vida"] == "Adolescencia" else "",
                "SI" if datos_usuario["curso_vida"] == "Juventud" else "",
                "SI" if datos_usuario["curso_vida"] == "Adultez" else "",
                "SI" if datos_usuario["curso_vida"] == "Vejez" else "",
                fecha_agudeza,
                fecha_planif,
                fecha_planif if fecha_planif else "",  # FECHA SUMINISTRO
                fecha_planif,  # SUMINISTRO DE METODO
                fecha_sangre,
                fecha_hdl,
                fecha_ldl,
                fecha_col_total,
                fecha_glicemia,
                fecha_creatinina,
                fecha_trigli
            ]
            ws.append(fila)
            row += 1

        if row > 2:
            self._aplicar_bordes(ws, 2, row - 1, len(encabezados))
        self._ajustar_columnas(ws)

    def _determinar_curso_vida(self, edad, unidad_edad) -> str:
        """Determina el curso de vida según Resolución 3280 de 2018"""
        if not edad or not unidad_edad:
            return "Desconocido"

        if unidad_edad == "3":  # Días
            edad_anos = edad / 365
        elif unidad_edad == "2":  # Meses
            edad_anos = edad / 12
        else:  # Años
            edad_anos = edad

        if edad_anos < 6:
            return "Primera Infancia"
        elif edad_anos < 12:
            return "Infancia"
        elif edad_anos < 18:
            return "Adolescencia"
        elif edad_anos < 29:
            return "Juventud"
        elif edad_anos < 60:
            return "Adultez"
        else:
            return "Vejez"
