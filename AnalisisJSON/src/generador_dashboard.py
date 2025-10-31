"""
Módulo para generar panel interactivo (dashboard) en Excel con gráficos y tablas dinámicas
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from typing import Dict, List, Any
from src.cargador_rips import CargadorRIPS
from collections import Counter


class GeneradorDashboard:
    """Genera dashboard interactivo con análisis visual de RIPS"""

    def __init__(self, cargador: CargadorRIPS):
        """
        Inicializa el generador de dashboard

        Args:
            cargador: Instancia de CargadorRIPS
        """
        self.cargador = cargador

        # Estilos
        self.title_font = Font(bold=True, size=16, color="FFFFFF")
        self.title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.subtitle_font = Font(bold=True, size=12, color="FFFFFF")
        self.subtitle_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, size=10, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.data_font = Font(size=10)
        self.number_font = Font(size=14, bold=True)
        self.center_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Side(style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        self.regimenes = {
            "01": "CONTRIBUTIVO", "02": "CONTRIBUTIVO", "03": "CONTRIBUTIVO",
            "04": "SUBSIDIADO", "05": "NO ASEGURADO",
            "06": "ESPECIAL", "07": "ESPECIAL", "08": "ESPECIAL", "09": "ESPECIAL"
        }

    def generar_dashboard(self, datos_consolidados: List[Dict[str, Any]], nombre_archivo: str):
        """
        Genera un dashboard interactivo en Excel

        Args:
            datos_consolidados: Lista con datos de todos los RIPS
            nombre_archivo: Nombre del archivo Excel a generar
        """
        wb = Workbook()
        wb.remove(wb.active)

        # Analizar datos
        estadisticas = self._analizar_datos(datos_consolidados)

        # Crear hojas del dashboard
        self._crear_resumen_ejecutivo(wb, estadisticas)
        self._crear_analisis_demografico(wb, estadisticas)
        self._crear_analisis_servicios(wb, estadisticas)
        self._crear_analisis_diagnosticos(wb, estadisticas)
        self._crear_analisis_procedimientos(wb, estadisticas)

        # Guardar
        wb.save(nombre_archivo)
        print(f"\nDashboard interactivo generado: {nombre_archivo}")

    def _analizar_datos(self, datos_consolidados: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analiza los datos consolidados y genera estadísticas"""
        total_usuarios = 0
        total_consultas = 0
        total_procedimientos = 0
        dist_sexo = Counter()
        dist_regimen = Counter()
        dist_edad = {"0-5": 0, "6-11": 0, "12-17": 0, "18-28": 0, "29-59": 0, "60+": 0}
        dist_municipios = Counter()
        diagnosticos = Counter()
        procedimientos = Counter()
        servicios_por_mes = Counter()

        for rips in datos_consolidados:
            datos = rips["datos"]
            usuarios = self.cargador.extraer_usuarios(datos)
            total_usuarios += len(usuarios)

            for usuario in usuarios:
                # Demografía
                sexo = usuario.get("codSexo", "")
                dist_sexo[sexo] += 1

                tipo_usuario = usuario.get("tipoUsuario", "")
                regimen = self.regimenes.get(tipo_usuario, "DESCONOCIDO")
                dist_regimen[regimen] += 1

                municipio = usuario.get("codMunicipioResidencia", "")
                dist_municipios[municipio] += 1

                # Edad
                fecha_nacimiento = usuario.get("fechaNacimiento", "")
                edad = self.cargador.calcular_edad(fecha_nacimiento)
                if edad is not None:
                    if edad < 6:
                        dist_edad["0-5"] += 1
                    elif edad < 12:
                        dist_edad["6-11"] += 1
                    elif edad < 18:
                        dist_edad["12-17"] += 1
                    elif edad < 29:
                        dist_edad["18-28"] += 1
                    elif edad < 60:
                        dist_edad["29-59"] += 1
                    else:
                        dist_edad["60+"] += 1

                # Consultas
                consultas = self.cargador.extraer_consultas(usuario)
                total_consultas += len(consultas)
                for consulta in consultas:
                    dx = consulta.get("codDiagnosticoPrincipal", "")
                    if dx:
                        diagnosticos[dx] += 1
                    fecha = consulta.get("fechaInicioAtencion", "")
                    if fecha:
                        mes = fecha[:7] if len(fecha) >= 7 else ""
                        if mes:
                            servicios_por_mes[mes] += 1

                # Procedimientos
                procs = self.cargador.extraer_procedimientos(usuario)
                total_procedimientos += len(procs)
                for proc in procs:
                    cod_proc = proc.get("codProcedimiento", "")
                    if cod_proc:
                        procedimientos[cod_proc] += 1
                    fecha = proc.get("fechaInicioAtencion", "")
                    if fecha:
                        mes = fecha[:7] if len(fecha) >= 7 else ""
                        if mes:
                            servicios_por_mes[mes] += 1

        return {
            "total_usuarios": total_usuarios,
            "total_consultas": total_consultas,
            "total_procedimientos": total_procedimientos,
            "dist_sexo": dict(dist_sexo),
            "dist_regimen": dict(dist_regimen),
            "dist_edad": dist_edad,
            "top_10_municipios": dict(dist_municipios.most_common(10)),
            "top_10_diagnosticos": dict(diagnosticos.most_common(10)),
            "top_10_procedimientos": dict(procedimientos.most_common(10)),
            "servicios_por_mes": dict(sorted(servicios_por_mes.items()))
        }

    def _crear_resumen_ejecutivo(self, wb: Workbook, stats: Dict[str, Any]):
        """Crea la hoja de resumen ejecutivo"""
        ws = wb.create_sheet("RESUMEN EJECUTIVO", 0)

        # Título principal
        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = "PANEL DE ANÁLISIS RIPS - RESUMEN EJECUTIVO"
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_alignment
        ws.row_dimensions[1].height = 30

        # Indicadores clave
        row = 3
        indicadores = [
            ("TOTAL USUARIOS", stats["total_usuarios"], "4472C4"),
            ("TOTAL CONSULTAS", stats["total_consultas"], "70AD47"),
            ("TOTAL PROCEDIMIENTOS", stats["total_procedimientos"], "FFC000"),
        ]

        col = 1
        for titulo, valor, color in indicadores:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            cell = ws.cell(row=row, column=col)
            cell.value = titulo
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = self.center_alignment
            cell.border = self.border

            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
            cell = ws.cell(row=row+1, column=col)
            cell.value = valor
            cell.font = Font(bold=True, size=20, color=color)
            cell.alignment = self.center_alignment
            cell.border = self.border
            ws.row_dimensions[row+1].height = 40

            col += 2

        # Distribución por sexo
        row = 6
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = "DISTRIBUCIÓN POR SEXO"
        cell.font = self.subtitle_font
        cell.fill = self.subtitle_fill
        cell.alignment = self.center_alignment

        row += 1
        ws.cell(row=row, column=1).value = "Sexo"
        ws.cell(row=row, column=2).value = "Cantidad"
        ws.cell(row=row, column=3).value = "Porcentaje"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        row += 1
        for sexo, cantidad in stats["dist_sexo"].items():
            porcentaje = (cantidad / stats["total_usuarios"] * 100) if stats["total_usuarios"] > 0 else 0
            ws.cell(row=row, column=1).value = "MASCULINO" if sexo == "M" else "FEMENINO" if sexo == "F" else sexo
            ws.cell(row=row, column=2).value = cantidad
            ws.cell(row=row, column=3).value = f"{porcentaje:.1f}%"
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1

        # Gráfico de torta - Sexo
        chart = PieChart()
        chart.title = "Distribución por Sexo"
        chart.style = 10
        labels = Reference(ws, min_col=1, min_row=8, max_row=7+len(stats["dist_sexo"]))
        data = Reference(ws, min_col=2, min_row=7, max_row=7+len(stats["dist_sexo"]))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 10
        chart.width = 15
        ws.add_chart(chart, "E6")

        # Distribución por régimen
        row += 2
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = "DISTRIBUCIÓN POR RÉGIMEN"
        cell.font = self.subtitle_font
        cell.fill = self.subtitle_fill
        cell.alignment = self.center_alignment

        row += 1
        ws.cell(row=row, column=1).value = "Régimen"
        ws.cell(row=row, column=2).value = "Cantidad"
        ws.cell(row=row, column=3).value = "Porcentaje"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        row += 1
        start_row_regimen = row
        for regimen, cantidad in stats["dist_regimen"].items():
            porcentaje = (cantidad / stats["total_usuarios"] * 100) if stats["total_usuarios"] > 0 else 0
            ws.cell(row=row, column=1).value = regimen
            ws.cell(row=row, column=2).value = cantidad
            ws.cell(row=row, column=3).value = f"{porcentaje:.1f}%"
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1

        # Gráfico de barras - Régimen
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Distribución por Régimen"
        chart.y_axis.title = "Cantidad de Usuarios"
        data = Reference(ws, min_col=2, min_row=start_row_regimen-1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=start_row_regimen, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 15
        ws.add_chart(chart, f"E{start_row_regimen-1}")

        # Ajustar columnas
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _crear_analisis_demografico(self, wb: Workbook, stats: Dict[str, Any]):
        """Crea la hoja de análisis demográfico"""
        ws = wb.create_sheet("ANÁLISIS DEMOGRÁFICO")

        # Título
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = "ANÁLISIS DEMOGRÁFICO"
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_alignment
        ws.row_dimensions[1].height = 30

        # Distribución por grupo etario
        row = 3
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = "DISTRIBUCIÓN POR GRUPO ETARIO"
        cell.font = self.subtitle_font
        cell.fill = self.subtitle_fill
        cell.alignment = self.center_alignment

        row += 1
        ws.cell(row=row, column=1).value = "Grupo de Edad"
        ws.cell(row=row, column=2).value = "Cantidad"
        ws.cell(row=row, column=3).value = "Porcentaje"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        row += 1
        start_row = row
        for grupo, cantidad in stats["dist_edad"].items():
            porcentaje = (cantidad / stats["total_usuarios"] * 100) if stats["total_usuarios"] > 0 else 0
            ws.cell(row=row, column=1).value = grupo
            ws.cell(row=row, column=2).value = cantidad
            ws.cell(row=row, column=3).value = f"{porcentaje:.1f}%"
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1

        # Gráfico de barras - Edad
        chart = BarChart()
        chart.type = "col"
        chart.style = 11
        chart.title = "Distribución por Grupos de Edad"
        chart.y_axis.title = "Cantidad de Usuarios"
        chart.x_axis.title = "Grupo Etario"
        data = Reference(ws, min_col=2, min_row=start_row-1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 18
        ws.add_chart(chart, "E3")

        # Top 10 municipios
        row += 2
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = "TOP 10 MUNICIPIOS"
        cell.font = self.subtitle_font
        cell.fill = self.subtitle_fill
        cell.alignment = self.center_alignment

        row += 1
        ws.cell(row=row, column=1).value = "Código Municipio"
        ws.cell(row=row, column=2).value = "Cantidad"
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        row += 1
        for municipio, cantidad in stats["top_10_municipios"].items():
            ws.cell(row=row, column=1).value = municipio
            ws.cell(row=row, column=2).value = cantidad
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = self.border
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _crear_analisis_servicios(self, wb: Workbook, stats: Dict[str, Any]):
        """Crea la hoja de análisis de servicios"""
        ws = wb.create_sheet("ANÁLISIS SERVICIOS")

        # Título
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = "ANÁLISIS DE SERVICIOS"
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_alignment
        ws.row_dimensions[1].height = 30

        # Tendencia de servicios por mes
        if stats["servicios_por_mes"]:
            row = 3
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws.cell(row=row, column=1)
            cell.value = "TENDENCIA DE SERVICIOS POR MES"
            cell.font = self.subtitle_font
            cell.fill = self.subtitle_fill
            cell.alignment = self.center_alignment

            row += 1
            ws.cell(row=row, column=1).value = "Mes"
            ws.cell(row=row, column=2).value = "Total Servicios"
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border

            row += 1
            start_row = row
            for mes, cantidad in stats["servicios_por_mes"].items():
                ws.cell(row=row, column=1).value = mes
                ws.cell(row=row, column=2).value = cantidad
                for col in range(1, 3):
                    ws.cell(row=row, column=col).border = self.border
                row += 1

            # Gráfico de líneas - Tendencia
            chart = LineChart()
            chart.title = "Tendencia de Servicios por Mes"
            chart.style = 12
            chart.y_axis.title = "Cantidad de Servicios"
            chart.x_axis.title = "Mes"
            data = Reference(ws, min_col=2, min_row=start_row-1, max_row=row-1)
            cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 20
            ws.add_chart(chart, "E3")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20

    def _crear_analisis_diagnosticos(self, wb: Workbook, stats: Dict[str, Any]):
        """Crea la hoja de análisis de diagnósticos"""
        ws = wb.create_sheet("TOP DIAGNÓSTICOS")

        # Título
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = "TOP 10 DIAGNÓSTICOS MÁS FRECUENTES"
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_alignment
        ws.row_dimensions[1].height = 30

        row = 3
        ws.cell(row=row, column=1).value = "Código CIE-10"
        ws.cell(row=row, column=2).value = "Nombre Diagnóstico"
        ws.cell(row=row, column=3).value = "Cantidad"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        row += 1
        start_row = row
        for codigo, cantidad in stats["top_10_diagnosticos"].items():
            info = self.cargador.obtener_info_diagnostico(codigo)
            ws.cell(row=row, column=1).value = codigo
            ws.cell(row=row, column=2).value = info.get("nombre", "")
            ws.cell(row=row, column=3).value = cantidad
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1

        # Gráfico de barras - Diagnósticos
        chart = BarChart()
        chart.type = "bar"
        chart.style = 13
        chart.title = "Top 10 Diagnósticos"
        chart.x_axis.title = "Cantidad"
        data = Reference(ws, min_col=3, min_row=start_row-1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 14
        chart.width = 20
        ws.add_chart(chart, "E3")

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15

    def _crear_analisis_procedimientos(self, wb: Workbook, stats: Dict[str, Any]):
        """Crea la hoja de análisis de procedimientos"""
        ws = wb.create_sheet("TOP PROCEDIMIENTOS")

        # Título
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = "TOP 10 PROCEDIMIENTOS MÁS FRECUENTES"
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_alignment
        ws.row_dimensions[1].height = 30

        row = 3
        ws.cell(row=row, column=1).value = "Código CUPS"
        ws.cell(row=row, column=2).value = "Nombre Procedimiento"
        ws.cell(row=row, column=3).value = "Cantidad"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        row += 1
        start_row = row
        for codigo, cantidad in stats["top_10_procedimientos"].items():
            nombre = self.cargador.obtener_nombre_cups(codigo)
            ws.cell(row=row, column=1).value = codigo
            ws.cell(row=row, column=2).value = nombre
            ws.cell(row=row, column=3).value = cantidad
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1

        # Gráfico de barras - Procedimientos
        chart = BarChart()
        chart.type = "bar"
        chart.style = 14
        chart.title = "Top 10 Procedimientos CUPS"
        chart.x_axis.title = "Cantidad"
        data = Reference(ws, min_col=3, min_row=start_row-1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 14
        chart.width = 20
        ws.add_chart(chart, "E3")

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 15
