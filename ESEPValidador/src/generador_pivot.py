import os
import sys
from pathlib import Path

# Agregar el directorio raíz del proyecto al sys.path
project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

import pandas as pd
from openpyxl import load_workbook
from config.config import OUTPUT_DIR

def generar_tabla_dinamica_errores():
    """
    Genera una tabla dinámica en una nueva hoja 'errores' para cada archivo Excel en OUTPUT_DIR,
    basada en la frecuencia de la columna 'codigo'.
    """
    if not os.path.exists(OUTPUT_DIR):
        print(f"Advertencia: La carpeta {OUTPUT_DIR} no existe.")
        return

    archivos_procesados = 0
    for archivo in os.listdir(OUTPUT_DIR):
        if archivo.lower().endswith('.xlsx'):
            ruta = os.path.join(OUTPUT_DIR, archivo)
            try:
                # Leer el archivo Excel, asumir primera hoja
                df = pd.read_excel(ruta, engine='openpyxl')
                
                if 'codigo' not in df.columns:
                    print(f"Advertencia: Columna 'codigo' no encontrada en {archivo}. Saltando.")
                    continue
                
                # Crear tabla de frecuencias
                pivot = df['codigo'].value_counts().reset_index()
                pivot.columns = ['codigo', 'frecuencia']
                
                # Cargar el workbook y agregar nueva hoja
                wb = load_workbook(ruta)
                if 'errores' in wb.sheetnames:
                    print(f"Advertencia: Hoja 'errores' ya existe en {archivo}. Sobrescribiendo.")
                    del wb['errores']
                
                ws = wb.create_sheet('errores')
                
                # Escribir headers
                ws['A1'] = 'codigo'
                ws['B1'] = 'frecuencia'
                
                # Escribir datos
                for idx, row in pivot.iterrows():
                    ws[f'A{idx+2}'] = row['codigo']
                    ws[f'B{idx+2}'] = row['frecuencia']
                
                wb.save(ruta)
                archivos_procesados += 1
                print(f"Tabla din&aacute;mica agregada a {archivo}")
                
            except Exception as e:
                print(f"Error procesando {archivo}: {e}")
    
    print(f"Procesamiento completado. Archivos procesados: {archivos_procesados}")

if __name__ == "__main__":
    generar_tabla_dinamica_errores()
